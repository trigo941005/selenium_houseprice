from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import time
from openpyxl import load_workbook
from openpyxl import Workbook

def title():
    title=["成交年月","地址","型態","總價","單價","建坪","地坪","樓別","屋齡","成交紀錄"]
    w2xlsx("HouseData.xlsx", title)
def w2xlsx(file, data):
    try:
        wb = load_workbook(filename=file)
        ws1 = wb["data"]
        ws1.append(data)
        wb.save(file)
    except FileNotFoundError:
        wb = Workbook()
        sheet = wb.create_sheet("data", 0)
        sheet.append(data)
        wb.save(file)
def search(last_data):
# 設置 Chrome Driver（假設你使用的是 Chrome 瀏覽器）
    global driver
    driver = webdriver.Chrome()
    # 開啟目標網站
    driver.get("https://price.houseprice.tw")  # 將此網址換成你的目標網站
    driver.maximize_window()
    # 使用 class 屬性來定位 <span> 元素
    span_element = driver.find_element(By.CLASS_NAME, "select_txt")
    time.sleep(2)
    span_element.click()
    time.sleep(2)
    span_element = driver.find_element(By.XPATH, "//span[text()='金門縣']")
    time.sleep(2)
    span_element.click()
    time.sleep(2)
    span_element = driver.find_element(By.XPATH, "//span[text()='金城鎮']")
    time.sleep(2)
    span_element.click()
    span_element = driver.find_element(By.XPATH, "//span[@data-v-95106727 and contains(@class, 'select_txt') and contains(@class, 'nowrap')]")
    time.sleep(2)
    span_element.click()
    span_element = driver.find_element(By.XPATH, "//span[text()='住宅大樓/華廈']")
    time.sleep(2)
    span_element.click()
    search_button = driver.find_element(By.XPATH, "//a[@class='search_btn']")
    time.sleep(2)
    search_button.click()

    wait = WebDriverWait(driver, 10)
    rows = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, "tr.group")))
    # 初始化空列表來存儲所有提取的資料
    all_data = []

    # 遍歷每個 <tr class="group"> 元素
    for row_element in rows:
        # 找到該 <tr> 標籤中的所有 <td> 標籤
        td_elements = row_element.find_elements(By.TAG_NAME, "td")

        # 初始化一個列表來存儲每個 row 的資料
        data_list = []

        # 遍歷每個 <td> 元素並提取內容
        for td in td_elements:
            try:
                # 如果有 <p> 標籤則提取其內容，否則提取 <td> 本身的內容
                if td.find_elements(By.TAG_NAME, "p"):
                    text_content = td.find_element(By.TAG_NAME, "p").text.strip()
                else:
                    text_content = td.text.strip()
            except NoSuchElementException:
                # 若沒有找到 <p> 標籤，則直接提取 <td> 的文本內容
                text_content = td.text.strip()
            
            # 將提取到的文本內容添加到列表中
            data_list.append(text_content)


        # 將該 row 的資料添加到 all_data 中
        all_data.append(data_list)

    # 輸出結果
    if last_data == all_data:
        return False
    for data in all_data:
        w2xlsx("HouseData.xlsx",data)
        NextPage()
    return all_data
def NextPage():
    global driver
    next_button = driver.find_element(By.XPATH, '//a[i[@class="sprite sprite-arrow_g r"]]')
    actions = ActionChains(driver)
    driver.execute_script("arguments[0].scrollIntoView(true);", next_button)
    next_button.click()
last_data = True
title()
while search(last_data)!=False:
    last_data = search(last_data)